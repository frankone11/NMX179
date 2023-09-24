import parametros
import requests
import xmltodict
from openpyxl import Workbook, load_workbook

def Inicia_Reporte(dia_ini:int, mes_ini:int, anio_ini:int, dia_fin:int, mes_fin:int, anio_fin:int):
	
	workbook = load_workbook(filename=parametros.archivo_entrada)
	sheet = workbook.active

	fecha_inicial = "{}-{:02d}-{:02d}".format(anio_ini, mes_ini, dia_ini)
	fecha_final = "{}-{:02d}-{:02d}".format(anio_fin, mes_fin, dia_fin)

	i = 2

	while True:
		numpozos = [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]
		volpozos = [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]
		cell = "A" + str(i)
		if sheet[cell].value == None:
			break

		rfc = sheet[cell].value

		j = 0
		while True:
			if j == 20:
				break
			numpozos[j] =  sheet.cell(row = i, column = j+2).value
			j += 1

		print(numpozos)

		j = 0
		while True:
			if j == 20 :
				break
			if numpozos[j] != None :
				nsm =  numpozos[j]
				formato = "Reporte de {} => Fecha inicial {} Fecha final: {}".format(rfc, fecha_inicial, fecha_final)
				print(formato)

				txml = '<soapenv:Envelope xmlns:soapenv="http://schemas.xmlsoap.org/soap/envelope/" xmlns:tem="http://tempuri.org/" xmlns:syc="http://schemas.datacontract.org/2004/07/SyC.CGRyF.WcfNMXService">\r\n'\
				'	<soapenv:Header/>\r\n'\
				'	<soapenv:Body>\r\n'\
				'		<tem:GetLecturasValidasRango>\r\n'\
				'			<!--Optional:-->\r\n'\
				'			<tem:filtro>\r\n'\
				'				<!--Optional:-->\r\n'\
				'				<syc:NSM>' + nsm + '</syc:NSM>\r\n'\
				'				<!--Optional:-->\r\n'\
				'				<syc:NSUE></syc:NSUE>\r\n'\
				'				<!--Optional:-->\r\n'\
				'				<syc:NSUT></syc:NSUT>\r\n'\
				'				<!--Optional:-->\r\n'\
				'				<syc:fechaFin>' + fecha_final +'</syc:fechaFin>\r\n'\
				'				<!--Optional:-->\r\n'\
				'				<syc:fechaInicio>' + fecha_inicial + '</syc:fechaInicio>\r\n'\
				'				<!--Optional:-->\r\n'\
				'				<syc:rfc>' + rfc + '</syc:rfc>\r\n'\
				'			</tem:filtro>\r\n'\
				'		</tem:GetLecturasValidasRango>\r\n'\
				'	</soapenv:Body>\r\n'\
				'</soapenv:Envelope>'
				headers = {'Content-Type': 'text/xml;charset=UTF-8',
					'SOAPAction' : 'http://tempuri.org/IServiceLecturas/GetLecturasValidasRango'}
				
				r = requests.post(parametros.URL, headers=headers, data=txml)

				rxml = r.text

				print("Resp: " + str(r.status_code))
				if r.status_code == 200:

					print("*****")

					stack = xmltodict.parse(rxml)

					elems = stack['s:Envelope']['s:Body']['GetLecturasValidasRangoResponse']['GetLecturasValidasRangoResult']['a:Datos']['a:LecValidas']

					wb = Workbook()

					ws = wb.active

					ws.cell(row=1, column=1).value = "RFC: " + rfc
					ws.append(["RFC", "NSM", "NSUE", "FECHA", "HORA", "LECTURA", "LATITUD", "LONGITUD", "CODIGO DE ERROR"])

					for elem in elems:
						tipomedicion = elem['a:tipoMedicion']
						r_rfc =  elem['a:RFC']
						r_nsm = elem['a:NSM']
						r_coderror = elem['a:codigoError']
						r_fecha = elem['a:fecha']
						r_hora = elem['a:hora']
						r_nsue = elem['a:NSUE']
						r_lectura = elem['a:lectura']
						r_latitud = elem['a:latitud']
						r_longitud = elem['a:longitud']

						format = "RFC: {} NSM: {} NSUE: {} Fecha: {} {} Lectura: {} Ubicación: [{}, {}] Código de Error: {}".format(
							r_rfc, r_nsm, r_nsue, r_fecha, r_hora, r_lectura, r_latitud, r_longitud, r_coderror
						)

						ws.append([r_rfc, r_nsm, r_nsue, r_fecha, r_hora, r_lectura, r_latitud, r_longitud, r_coderror])
						
						print(format)
						print("*****")

					wb.save("{}_{}.xlsx".format(rfc,nsm))
					print("Fin de impresión")
				else:
					print("Error de respuesta del servidor")
			else:
				print("Celda en blanco")

